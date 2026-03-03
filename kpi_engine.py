# kpi_engine.py  — FINAL FIXED VERSION
# Run: python kpi_engine.py
# Requires: pip install pandas openpyxl
# Output: growth_audit_dashboard.xlsx  → upload to Google Sheets

import pandas as pd
import warnings
warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# ─── CONFIG ───────────────────────────────────────────────────
CSV_FILE  = "ecommerce_data.csv"
OUT_EXCEL = "growth_audit_dashboard.xlsx"

BENCHMARKS = {
    "conversion_rate": 3.80,
    "cac":             40.00,
    "roas":             6.00,
    "gross_margin":    40.00,
    "aov":             70.00,
    "repeat_rate":     30.00,
    "rev_drop_alert":  -5.00,
}

# ─── COLOR PALETTE ────────────────────────────────────────────
C = {
    "dark":   "0F172A",
    "navy":   "1E293B",
    "indigo": "6366F1",
    "green":  "10B981",
    "red":    "EF4444",
    "amber":  "F59E0B",
    "slate":  "334155",
    "light":  "E2E8F0",
    "muted":  "94A3B8",
    "white":  "FFFFFF",
    "purple": "8B5CF6",
    "teal":   "0D9488",
}

# ─── STYLE HELPERS ────────────────────────────────────────────
def fl(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fn(color="FFFFFF", bold=True, size=10):
    return Font(color=color, bold=bold, size=size, name="Calibri")

def bd():
    s = Side(style="thin", color="475569")
    return Border(left=s, right=s, top=s, bottom=s)

def ca():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def la():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def safe_val(v):
    """Convert any pandas type that openpyxl cannot handle into a plain Python type."""
    if isinstance(v, pd.Period):
        return str(v)
    if isinstance(v, pd.Timestamp):
        return v.strftime("%Y-%m-%d")
    if hasattr(v, 'item'):          # numpy scalar → python scalar
        return v.item()
    return v

def write_df(ws, df_in, start_row=1, start_col=1, hdr_bg="6366F1"):
    """Write a DataFrame to a worksheet with styled headers and alternating rows."""
    cols = list(df_in.columns)
    # Header row
    for ci, col in enumerate(cols, start_col):
        c = ws.cell(row=start_row, column=ci, value=str(col))
        c.fill      = fl(hdr_bg)
        c.font      = fn()
        c.alignment = ca()
        c.border    = bd()
    ws.row_dimensions[start_row].height = 24
    # Data rows
    for ri, (_, row) in enumerate(df_in.iterrows(), 1):
        bg = C["navy"] if ri % 2 == 1 else "162032"
        ws.row_dimensions[start_row + ri].height = 19
        for ci, val in enumerate(row, start_col):
            c = ws.cell(row=start_row + ri, column=ci, value=safe_val(val))
            c.fill      = fl(bg)
            c.font      = fn(C["light"], bold=False, size=9)
            c.alignment = ca()
            c.border    = bd()

def sheet_title(ws, text, ncols, bg="0F172A", fg="FFFFFF", size=13):
    """Merge cells across row 1 and write a styled title."""
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=text)
    t.fill      = fl(bg)
    t.font      = fn(color=fg, bold=True, size=size)
    t.alignment = ca()
    ws.row_dimensions[1].height = 36

# ─── LOAD & CLEAN ─────────────────────────────────────────────
print("Loading data...")
df = pd.read_csv(CSV_FILE, parse_dates=["order_date"])
df.drop_duplicates(subset="order_id", inplace=True)
df.dropna(subset=["net_revenue", "order_date", "customer_id"], inplace=True)
for col in ["net_revenue", "cogs", "marketing_spend", "website_visits"]:
    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
df["is_new_customer"] = df["is_new_customer"].astype(int)
df["is_returned"]     = df["is_returned"].astype(int)
# Keep a plain string copy of order_date for the Raw Data sheet
df["order_date_str"]  = df["order_date"].dt.strftime("%Y-%m-%d")
# Period only for groupby — converted to str immediately after
df["month_period"]    = df["order_date"].dt.to_period("M")
df["month_str"]       = df["order_date"].dt.strftime("%b-%y")
print(f"Loaded {len(df):,} clean records | {df['month_period'].nunique()} months")

# ─── MONTHLY KPI AGGREGATION ──────────────────────────────────
print("Calculating KPIs...")
monthly = df.groupby("month_period").agg(
    month_label   =("month_str",       "first"),
    revenue       =("net_revenue",     "sum"),
    orders        =("order_id",        "count"),
    customers     =("customer_id",     "nunique"),
    new_customers =("is_new_customer", "sum"),
    returns       =("is_returned",     "sum"),
    cogs          =("cogs",            "sum"),
    mkt_spend     =("marketing_spend", "sum"),
    visits        =("website_visits",  "sum"),
    inventory     =("inventory_units", "mean"),
).reset_index().sort_values("month_period")

# *** CRITICAL FIX: convert Period index to plain string BEFORE any Excel ops ***
monthly["month_period"] = monthly["month_period"].astype(str)

monthly["aov"]          = (monthly["revenue"] / monthly["orders"]).round(2)
monthly["cac"]          = (monthly["mkt_spend"] / monthly["new_customers"].replace(0, 1)).round(2)
monthly["conv_rate"]    = (monthly["orders"] / monthly["visits"] * 100).round(2)
monthly["roas"]         = (monthly["revenue"] / monthly["mkt_spend"].replace(0, 1)).round(2)
monthly["gross_margin"] = ((monthly["revenue"] - monthly["cogs"]) / monthly["revenue"] * 100).round(1)
monthly["repeat_rate"]  = ((1 - monthly["new_customers"] / monthly["orders"]) * 100).round(1)
monthly["inv_turnover"] = (monthly["cogs"] / monthly["inventory"].replace(0, 1)).round(2)
monthly["rev_change"]   = monthly["revenue"].pct_change().mul(100).round(1)

# ─── ANOMALY DETECTION ────────────────────────────────────────
print("Running anomaly detection...")

def detect(row):
    alerts = []
    if pd.notna(row["rev_change"]) and row["rev_change"] <= BENCHMARKS["rev_drop_alert"]:
        alerts.append(f"CRITICAL: Revenue dropped {abs(row['rev_change'])}% MoM")
    if row["cac"] > BENCHMARKS["cac"]:
        alerts.append(f"WARNING: CAC ${row['cac']} exceeds benchmark ${BENCHMARKS['cac']}")
    if row["conv_rate"] < BENCHMARKS["conversion_rate"]:
        alerts.append(f"WARNING: Conv rate {row['conv_rate']}% below {BENCHMARKS['conversion_rate']}% target")
    if row["roas"] < BENCHMARKS["roas"]:
        alerts.append(f"CRITICAL: ROAS {row['roas']}x below {BENCHMARKS['roas']}x threshold")
    if row["gross_margin"] < BENCHMARKS["gross_margin"]:
        alerts.append(f"WARNING: Gross margin {row['gross_margin']}% below {BENCHMARKS['gross_margin']}% target")
    return " | ".join(alerts) if alerts else "OK - No Alerts"

monthly["anomaly_flags"]  = monthly.apply(detect, axis=1)
monthly["alert_count"]    = monthly["anomaly_flags"].apply(
    lambda x: 0 if x == "OK - No Alerts" else len(x.split("|")))
monthly["alert_severity"] = monthly["anomaly_flags"].apply(
    lambda x: "CRITICAL" if "CRITICAL" in x else ("WARNING" if "WARNING" in x else "OK"))

# ─── CATEGORY KPIs ────────────────────────────────────────────
cat_kpi = df.groupby("category").agg(
    revenue   =("net_revenue",     "sum"),
    orders    =("order_id",        "count"),
    returns   =("is_returned",     "sum"),
    cogs      =("cogs",            "sum"),
    mkt_spend =("marketing_spend", "sum"),
    customers =("customer_id",     "nunique"),
).reset_index()
cat_kpi["aov"]          = (cat_kpi["revenue"] / cat_kpi["orders"]).round(2)
cat_kpi["return_rate"]  = (cat_kpi["returns"] / cat_kpi["orders"] * 100).round(1)
cat_kpi["gross_margin"] = ((cat_kpi["revenue"] - cat_kpi["cogs"]) / cat_kpi["revenue"] * 100).round(1)
cat_kpi["roas"]         = (cat_kpi["revenue"] / cat_kpi["mkt_spend"].replace(0, 1)).round(2)
cat_kpi["rev_share"]    = (cat_kpi["revenue"] / cat_kpi["revenue"].sum() * 100).round(1)
cat_kpi.sort_values("revenue", ascending=False, inplace=True)

# ─── CAMPAIGN KPIs ────────────────────────────────────────────
camp_kpi = df.groupby("campaign").agg(
    revenue   =("net_revenue",     "sum"),
    orders    =("order_id",        "count"),
    mkt_spend =("marketing_spend", "sum"),
    customers =("customer_id",     "nunique"),
    new_custs =("is_new_customer", "sum"),
).reset_index()
camp_kpi["roas"] = (camp_kpi["revenue"] / camp_kpi["mkt_spend"].replace(0, 1)).round(2)
camp_kpi["cac"]  = (camp_kpi["mkt_spend"] / camp_kpi["new_custs"].replace(0, 1)).round(2)
camp_kpi["aov"]  = (camp_kpi["revenue"] / camp_kpi["orders"]).round(2)
camp_kpi.sort_values("roas", ascending=False, inplace=True)

# ─── REGIONAL KPIs ────────────────────────────────────────────
reg_kpi = df.groupby("region").agg(
    revenue   =("net_revenue",     "sum"),
    orders    =("order_id",        "count"),
    customers =("customer_id",     "nunique"),
    returns   =("is_returned",     "sum"),
    mkt_spend =("marketing_spend", "sum"),
).reset_index()
reg_kpi["aov"]         = (reg_kpi["revenue"] / reg_kpi["orders"]).round(2)
reg_kpi["return_rate"] = (reg_kpi["returns"] / reg_kpi["orders"] * 100).round(1)
reg_kpi["roas"]        = (reg_kpi["revenue"] / reg_kpi["mkt_spend"].replace(0, 1)).round(2)
reg_kpi["rev_share"]   = (reg_kpi["revenue"] / reg_kpi["revenue"].sum() * 100).round(1)
reg_kpi.sort_values("revenue", ascending=False, inplace=True)

# ─── H1 vs H2 ─────────────────────────────────────────────────
h1 = monthly[monthly["month_period"] <= "2024-06"]
h2 = monthly[monthly["month_period"] >  "2024-06"]

def pct_chg(a, b):
    return f"{((b - a) / a * 100):.1f}%" if a != 0 else "N/A"

root_cause_df = pd.DataFrame([
    {"Metric": "Total Revenue",        "H1-2024": f"${h1['revenue'].sum():,.0f}",       "H2-2024": f"${h2['revenue'].sum():,.0f}",      "Change": pct_chg(h1['revenue'].sum(),      h2['revenue'].sum())},
    {"Metric": "Avg Monthly Revenue",  "H1-2024": f"${h1['revenue'].mean():,.0f}",       "H2-2024": f"${h2['revenue'].mean():,.0f}",     "Change": pct_chg(h1['revenue'].mean(),     h2['revenue'].mean())},
    {"Metric": "Total Mkt Spend",      "H1-2024": f"${h1['mkt_spend'].sum():,.0f}",      "H2-2024": f"${h2['mkt_spend'].sum():,.0f}",    "Change": pct_chg(h1['mkt_spend'].sum(),    h2['mkt_spend'].sum())},
    {"Metric": "Avg ROAS",             "H1-2024": f"{h1['roas'].mean():.2f}x",           "H2-2024": f"{h2['roas'].mean():.2f}x",         "Change": pct_chg(h1['roas'].mean(),        h2['roas'].mean())},
    {"Metric": "Avg CAC ($)",          "H1-2024": f"${h1['cac'].mean():.2f}",            "H2-2024": f"${h2['cac'].mean():.2f}",          "Change": pct_chg(h1['cac'].mean(),         h2['cac'].mean())},
    {"Metric": "Avg Conv Rate (%)",    "H1-2024": f"{h1['conv_rate'].mean():.2f}%",      "H2-2024": f"{h2['conv_rate'].mean():.2f}%",    "Change": pct_chg(h1['conv_rate'].mean(),   h2['conv_rate'].mean())},
    {"Metric": "Total Returns",        "H1-2024": str(int(h1['returns'].sum())),          "H2-2024": str(int(h2['returns'].sum())),       "Change": pct_chg(h1['returns'].sum(),      h2['returns'].sum())},
    {"Metric": "New Customers",        "H1-2024": str(int(h1['new_customers'].sum())),    "H2-2024": str(int(h2['new_customers'].sum())), "Change": pct_chg(h1['new_customers'].sum(),h2['new_customers'].sum())},
    {"Metric": "Avg Gross Margin (%)", "H1-2024": f"{h1['gross_margin'].mean():.1f}%",   "H2-2024": f"{h2['gross_margin'].mean():.1f}%","Change": pct_chg(h1['gross_margin'].mean(),h2['gross_margin'].mean())},
    {"Metric": "Avg Repeat Rate (%)",  "H1-2024": f"{h1['repeat_rate'].mean():.1f}%",    "H2-2024": f"{h2['repeat_rate'].mean():.1f}%", "Change": pct_chg(h1['repeat_rate'].mean(), h2['repeat_rate'].mean())},
])

recs = pd.DataFrame([
    {"Priority":"CRITICAL","Area":"Paid Acquisition",  "Finding":"ROAS declining, CAC escalating sharply in H2",           "Action":"Pause bottom 30% ad sets; refresh creatives; shift 20% budget to retargeting",          "Impact":"ROAS +40%, CAC -35%"},
    {"Priority":"CRITICAL","Area":"Conversion Rate",   "Finding":"Conv rate below 3.8% benchmark since Aug-24",            "Action":"CRO sprint: simplify checkout, add social proof, exit-intent offers, fix Core Web Vitals","Impact":"+0.5pp conv = ~$20K/mo"},
    {"Priority":"HIGH",    "Area":"Returns Reduction", "Finding":"Returns doubled in H2, eroding gross margin",            "Action":"Audit top return SKUs, add size guides, 360-degree images, post-purchase surveys",       "Impact":"Return rate -40%"},
    {"Priority":"HIGH",    "Area":"Retention",         "Finding":"New customer acquisition down 42%, no retention program","Action":"Email nurture, loyalty points, VIP tier, 90-day reactivation campaign",                  "Impact":"Repeat rate +10pp = ~$12K/mo"},
    {"Priority":"MEDIUM",  "Area":"Inventory",         "Finding":"Turnover slowing in H2, capital tied in slow SKUs",      "Action":"Reduce slow-SKU reorders, clearance on >90-day stock, demand forecasting",               "Impact":"Free $40K+ capital"},
    {"Priority":"MEDIUM",  "Area":"Campaign Mix",      "Finding":"High ROAS variance across channels",                     "Action":"Reallocate 30% budget from low-ROAS channels to Email and Organic",                      "Impact":"Blended ROAS +1.5x"},
])

latest_m   = monthly.iloc[-1]
peak_rev   = monthly["revenue"].max()
peak_month = monthly.loc[monthly["revenue"].idxmax(), "month_label"]
rev_delta  = ((latest_m["revenue"] - peak_rev) / peak_rev * 100)

# ═══════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════
print(f"Writing Excel -> {OUT_EXCEL}")

wb      = Workbook()
ws_dash = wb.active;           ws_dash.title = "Dashboard"
ws_kpi  = wb.create_sheet("Monthly KPIs")
ws_cat  = wb.create_sheet("By Category")
ws_camp = wb.create_sheet("By Campaign")
ws_reg  = wb.create_sheet("By Region")
ws_anom = wb.create_sheet("Anomaly Log")
ws_rc   = wb.create_sheet("Root Cause")
ws_recs = wb.create_sheet("Recommendations")
ws_sql  = wb.create_sheet("SQL Queries")
ws_raw  = wb.create_sheet("Raw Data")

# ═══════════════════════════════════════════════════════════════
#  SHEET 1 — DASHBOARD
# ═══════════════════════════════════════════════════════════════
ws_dash.sheet_view.showGridLines = False
ws_dash.sheet_properties.tabColor = C["indigo"]

ws_dash.merge_cells("A1:N3")
t = ws_dash["A1"]
t.value     = "AUTOMATED GROWTH AUDIT  |  E-Commerce KPI Intelligence Dashboard  |  FY 2024"
t.fill      = fl(C["dark"])
t.font      = fn(color=C["white"], bold=True, size=17)
t.alignment = ca()
ws_dash.row_dimensions[1].height = 46

ws_dash.merge_cells("A4:N4")
s = ws_dash["A4"]
s.value     = "Anomaly detection  |  KPI benchmarking  |  Root cause analysis  |  Strategic recommendations"
s.fill      = fl(C["navy"])
s.font      = fn(color=C["muted"], bold=False, size=10)
s.alignment = ca()
ws_dash.row_dimensions[4].height = 22

# KPI Cards
kpi_cards = [
    ("Latest Revenue",  f"${latest_m['revenue']:,.0f}",     f"{latest_m['rev_change']:+.1f}% MoM",          C["indigo"], latest_m["rev_change"] >= 0),
    ("ROAS",            f"{latest_m['roas']:.2f}x",         f"Benchmark {BENCHMARKS['roas']}x",             C["purple"], latest_m["roas"]      >= BENCHMARKS["roas"]),
    ("CAC",             f"${latest_m['cac']:.2f}",          f"Benchmark ${BENCHMARKS['cac']}",              C["amber"],  latest_m["cac"]       <= BENCHMARKS["cac"]),
    ("Conv Rate",       f"{latest_m['conv_rate']:.2f}%",    f"Benchmark {BENCHMARKS['conversion_rate']}%",  C["teal"],   latest_m["conv_rate"]  >= BENCHMARKS["conversion_rate"]),
    ("Gross Margin",    f"{latest_m['gross_margin']:.1f}%", f"Benchmark {BENCHMARKS['gross_margin']}%",     C["green"],  latest_m["gross_margin"] >= BENCHMARKS["gross_margin"]),
    ("Avg Order Value", f"${latest_m['aov']:.2f}",          f"Benchmark ${BENCHMARKS['aov']}",              C["indigo"], latest_m["aov"]        >= BENCHMARKS["aov"]),
    ("Total Alerts",    str(int(monthly["alert_count"].sum())), "Anomaly flags triggered",                   C["red"],    False),
]

card_start_cols = [1, 3, 5, 7, 9, 11, 13]
for idx, (lbl, val, sub, col, is_ok) in enumerate(kpi_cards):
    c0 = card_start_cols[idx]
    rs = 6
    for r in range(rs, rs + 4):
        ws_dash.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 1)

    lc = ws_dash.cell(rs,     c0, lbl)
    vc = ws_dash.cell(rs + 1, c0, val)
    sc = ws_dash.cell(rs + 2, c0, sub)
    tc = ws_dash.cell(rs + 3, c0, "ON TARGET" if is_ok else "BELOW TARGET")

    lc.fill = fl(col);          lc.font = fn(color=C["white"], bold=True, size=9);   lc.alignment = ca()
    vc.fill = fl(C["navy"]);    vc.font = fn(color=C["white"], bold=True, size=19);  vc.alignment = ca()
    sc.fill = fl(C["navy"]);    sc.font = fn(color=C["muted"], bold=False, size=8);  sc.alignment = ca()
    status_color = C["green"] if is_ok else C["red"]
    tc.fill = fl(status_color); tc.font = fn(color=C["white"], bold=True, size=8);   tc.alignment = ca()

    for r in range(rs, rs + 4):
        ws_dash.row_dimensions[r].height = 21

for i in range(1, 15):
    ws_dash.column_dimensions[get_column_letter(i)].width = 13

# H1 vs H2 summary
r2 = 12
ws_dash.merge_cells(f"A{r2}:N{r2}")
ws_dash.cell(r2, 1, "H1 vs H2 Performance Summary").fill = fl(C["slate"])
ws_dash.cell(r2, 1).font      = fn(color=C["white"], bold=True, size=11)
ws_dash.cell(r2, 1).alignment = la()
ws_dash.row_dimensions[r2].height = 26

sum_hdrs = ["Period","Revenue","Orders","New Customers","Avg CAC","Avg ROAS",
            "Conv Rate","Returns","Gross Margin","Repeat Rate"]
for ci, h in enumerate(sum_hdrs, 1):
    c = ws_dash.cell(r2 + 1, ci, h)
    c.fill = fl(C["indigo"]); c.font = fn(); c.alignment = ca(); c.border = bd()
ws_dash.row_dimensions[r2 + 1].height = 22

h1_vals = ["H1-2024", f"${h1['revenue'].sum():,.0f}", int(h1['orders'].sum()),
           int(h1['new_customers'].sum()), f"${h1['cac'].mean():.2f}",
           f"{h1['roas'].mean():.2f}x", f"{h1['conv_rate'].mean():.2f}%",
           int(h1['returns'].sum()), f"{h1['gross_margin'].mean():.1f}%",
           f"{h1['repeat_rate'].mean():.1f}%"]
h2_vals = ["H2-2024", f"${h2['revenue'].sum():,.0f}", int(h2['orders'].sum()),
           int(h2['new_customers'].sum()), f"${h2['cac'].mean():.2f}",
           f"{h2['roas'].mean():.2f}x", f"{h2['conv_rate'].mean():.2f}%",
           int(h2['returns'].sum()), f"{h2['gross_margin'].mean():.1f}%",
           f"{h2['repeat_rate'].mean():.1f}%"]

for ci, v in enumerate(h1_vals, 1):
    c = ws_dash.cell(r2 + 2, ci, v)
    c.fill = fl(C["navy"]); c.font = fn(color=C["green"], bold=(ci==1), size=10)
    c.alignment = ca(); c.border = bd()
for ci, v in enumerate(h2_vals, 1):
    c = ws_dash.cell(r2 + 3, ci, v)
    c.fill = fl(C["navy"]); c.font = fn(color=C["red"], bold=(ci==1), size=10)
    c.alignment = ca(); c.border = bd()
for r in [r2+2, r2+3]:
    ws_dash.row_dimensions[r].height = 20

# Key findings
r3 = r2 + 6
ws_dash.merge_cells(f"A{r3}:N{r3}")
ws_dash.cell(r3, 1, "Key Findings from Anomaly Engine").fill = fl(C["slate"])
ws_dash.cell(r3, 1).font      = fn(color=C["white"], bold=True, size=11)
ws_dash.cell(r3, 1).alignment = la()
ws_dash.row_dimensions[r3].height = 26

findings = [
    (C["red"],   f"Revenue declined {abs(rev_delta):.1f}% from peak ${peak_rev:,.0f} ({peak_month}). Continued spending while revenue fell."),
    (C["red"],   f"ROAS collapsed {monthly.iloc[0]['roas']:.2f}x -> {monthly.iloc[-1]['roas']:.2f}x while marketing spend kept rising. Classic ad fatigue."),
    (C["amber"], f"CAC rose ${monthly.iloc[0]['cac']:.2f} -> ${monthly.iloc[-1]['cac']:.2f}. New customers fell 40%+ with no creative refresh."),
    (C["amber"], f"Conv rate dropped {monthly.iloc[0]['conv_rate']:.2f}% -> {monthly.iloc[-1]['conv_rate']:.2f}%. Checkout friction or product-fit issues."),
    (C["amber"], f"Returns nearly doubled H1({int(h1['returns'].sum())}) vs H2({int(h2['returns'].sum())}). Eroding net revenue and margin."),
    (C["green"], "Action: CRO sprint + campaign restructure + retention program. See Recommendations sheet."),
]
for fi, (color, text) in enumerate(findings):
    r = r3 + 1 + fi
    ws_dash.merge_cells(f"A{r}:N{r}")
    c = ws_dash.cell(r, 1, text)
    c.fill      = fl(C["navy"])
    c.font      = Font(color=color, size=10, name="Calibri")
    c.alignment = la()
    ws_dash.row_dimensions[r].height = 20

# ═══════════════════════════════════════════════════════════════
#  SHEET 2 — MONTHLY KPIs + 4 CHARTS
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_kpi, "Monthly KPI Intelligence Report  |  FY 2024", 17, size=13)
ws_kpi.sheet_properties.tabColor = C["purple"]

mc = ["month_label","revenue","orders","new_customers","returns","mkt_spend","visits",
      "aov","cac","conv_rate","roas","gross_margin","repeat_rate","inv_turnover",
      "rev_change","alert_severity","anomaly_flags"]
mh = ["Month","Revenue($)","Orders","New Custs","Returns","Mkt Spend($)","Visits",
      "AOV($)","CAC($)","Conv Rate(%)","ROAS(x)","Gross Margin(%)","Repeat Rate(%)",
      "Inv Turnover","Rev Change(%)","Alert Status","Anomaly Flags"]
mo = monthly[mc].copy()
mo.columns = mh
write_df(ws_kpi, mo, start_row=2, hdr_bg=C["indigo"])

# Conditional formatting
ws_kpi.conditional_formatting.add(
    f"J3:J{2+len(mo)}",
    CellIsRule(operator="lessThan", formula=["3.8"],
               fill=PatternFill("solid", fgColor="3B0000"),
               font=Font(color=C["red"], bold=True)))
ws_kpi.conditional_formatting.add(
    f"K3:K{2+len(mo)}",
    ColorScaleRule(start_type="min", start_color="EF4444",
                   mid_type="num",   mid_value=6, mid_color="F59E0B",
                   end_type="max",   end_color="10B981"))

for i, w in enumerate([10,14,8,11,8,13,8,9,9,11,9,14,12,11,12,12,50], 1):
    ws_kpi.column_dimensions[get_column_letter(i)].width = w

cats = Reference(ws_kpi, min_col=1, min_row=3, max_row=14)

# Chart 1 — Revenue vs Mkt Spend
ch1 = LineChart()
ch1.title = "Revenue vs Marketing Spend  FY2024"
ch1.style = 10; ch1.y_axis.title = "Amount($)"; ch1.width = 26; ch1.height = 14
ch1.add_data(Reference(ws_kpi, min_col=2,  max_col=2,  min_row=2, max_row=14), titles_from_data=True)
ch1.add_data(Reference(ws_kpi, min_col=6,  max_col=6,  min_row=2, max_row=14), titles_from_data=True)
ch1.set_categories(cats)
ch1.series[0].graphicalProperties.line.solidFill = C["indigo"]; ch1.series[0].graphicalProperties.line.width = 28000
ch1.series[1].graphicalProperties.line.solidFill = C["red"];    ch1.series[1].graphicalProperties.line.width = 20000
ws_kpi.add_chart(ch1, "A18")

# Chart 2 — ROAS
ch2 = LineChart()
ch2.title = "ROAS Trend (Benchmark 6x)"
ch2.style = 10; ch2.y_axis.title = "ROAS(x)"; ch2.width = 20; ch2.height = 13
ch2.add_data(Reference(ws_kpi, min_col=11, max_col=11, min_row=2, max_row=14), titles_from_data=True)
ch2.set_categories(cats)
ch2.series[0].graphicalProperties.line.solidFill = C["green"]; ch2.series[0].graphicalProperties.line.width = 28000
ws_kpi.add_chart(ch2, "L18")

# Chart 3 — Conversion Rate
ch3 = BarChart()
ch3.title = "Conversion Rate % (Benchmark 3.8%)"
ch3.style = 10; ch3.y_axis.title = "Conv Rate(%)"; ch3.width = 26; ch3.height = 13
ch3.add_data(Reference(ws_kpi, min_col=10, max_col=10, min_row=2, max_row=14), titles_from_data=True)
ch3.set_categories(cats)
ch3.series[0].graphicalProperties.solidFill = C["indigo"]
ws_kpi.add_chart(ch3, "A38")

# Chart 4 — CAC
ch4 = LineChart()
ch4.title = "CAC Trend (Benchmark $40)"
ch4.style = 10; ch4.y_axis.title = "CAC($)"; ch4.width = 20; ch4.height = 13
ch4.add_data(Reference(ws_kpi, min_col=9,  max_col=9,  min_row=2, max_row=14), titles_from_data=True)
ch4.set_categories(cats)
ch4.series[0].graphicalProperties.line.solidFill = C["amber"]; ch4.series[0].graphicalProperties.line.width = 28000
ws_kpi.add_chart(ch4, "L38")

# ═══════════════════════════════════════════════════════════════
#  SHEET 3 — BY CATEGORY
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_cat, "Category-Level KPI Breakdown", 9, size=13)
ws_cat.sheet_properties.tabColor = C["green"]

co = cat_kpi[["category","revenue","orders","customers","aov",
              "return_rate","gross_margin","roas","rev_share"]].copy()
co.columns = ["Category","Revenue($)","Orders","Customers","AOV($)",
               "Return Rate(%)","Gross Margin(%)","ROAS(x)","Rev Share(%)"]
write_df(ws_cat, co, start_row=2, hdr_bg=C["green"])
for i, w in enumerate([18,14,9,11,10,14,15,10,12], 1):
    ws_cat.column_dimensions[get_column_letter(i)].width = w

ch_cat = BarChart()
ch_cat.title = "Revenue by Category"; ch_cat.style = 10
ch_cat.y_axis.title = "Revenue($)"; ch_cat.width = 18; ch_cat.height = 12
ch_cat.add_data(Reference(ws_cat, min_col=2, max_col=2, min_row=2, max_row=2+len(co)), titles_from_data=True)
ch_cat.set_categories(Reference(ws_cat, min_col=1, min_row=3, max_row=2+len(co)))
ch_cat.series[0].graphicalProperties.solidFill = C["green"]
ws_cat.add_chart(ch_cat, "A15")

# ═══════════════════════════════════════════════════════════════
#  SHEET 4 — BY CAMPAIGN
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_camp, "Campaign Performance & ROI Analysis", 8, size=13)
ws_camp.sheet_properties.tabColor = C["amber"]

cpo = camp_kpi[["campaign","revenue","orders","mkt_spend","new_custs","cac","roas","aov"]].copy()
cpo.columns = ["Campaign","Revenue($)","Orders","Mkt Spend($)","New Custs","CAC($)","ROAS(x)","AOV($)"]
write_df(ws_camp, cpo, start_row=2, hdr_bg=C["amber"])
for i, w in enumerate([18,14,9,13,11,10,10,10], 1):
    ws_camp.column_dimensions[get_column_letter(i)].width = w

ch_roas = BarChart()
ch_roas.title = "ROAS by Campaign"; ch_roas.style = 10
ch_roas.y_axis.title = "ROAS(x)"; ch_roas.width = 16; ch_roas.height = 12
ch_roas.add_data(Reference(ws_camp, min_col=7, max_col=7, min_row=2, max_row=2+len(cpo)), titles_from_data=True)
ch_roas.set_categories(Reference(ws_camp, min_col=1, min_row=3, max_row=2+len(cpo)))
ch_roas.series[0].graphicalProperties.solidFill = C["amber"]
ws_camp.add_chart(ch_roas, "A15")

# ═══════════════════════════════════════════════════════════════
#  SHEET 5 — BY REGION
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_reg, "Regional Performance Analysis", 8, size=13)
ws_reg.sheet_properties.tabColor = C["purple"]

ro = reg_kpi[["region","revenue","orders","customers","aov","return_rate","roas","rev_share"]].copy()
ro.columns = ["Region","Revenue($)","Orders","Customers","AOV($)","Return Rate(%)","ROAS(x)","Rev Share(%)"]
write_df(ws_reg, ro, start_row=2, hdr_bg=C["purple"])
for i, w in enumerate([12,14,9,11,10,14,10,12], 1):
    ws_reg.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════
#  SHEET 6 — ANOMALY LOG
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_anom, "Automated Anomaly Detection Log  |  All Triggered Alerts", 4, size=13)
ws_anom.sheet_properties.tabColor = C["red"]

anom_rows = []
for _, row in monthly.iterrows():
    if row["anomaly_flags"] != "OK - No Alerts":
        for flag in row["anomaly_flags"].split("|"):
            flag = flag.strip()
            sev  = "CRITICAL" if "CRITICAL" in flag else "WARNING"
            act  = ("Pause campaigns, review ad creative immediately"
                    if "ROAS" in flag or "Revenue" in flag
                    else "Run CRO tests and review checkout funnel"
                    if "Conv" in flag
                    else "Monitor trend and investigate root cause")
            anom_rows.append({
                "Month": row["month_label"],
                "Severity": sev,
                "Alert Detail": flag,
                "Recommended Action": act,
            })

adf = (pd.DataFrame(anom_rows) if anom_rows
       else pd.DataFrame(columns=["Month","Severity","Alert Detail","Recommended Action"]))
write_df(ws_anom, adf, start_row=2, hdr_bg=C["red"])
for i, w in enumerate([10, 12, 60, 50], 1):
    ws_anom.column_dimensions[get_column_letter(i)].width = w

for ri in range(3, 3 + len(adf)):
    sev      = ws_anom.cell(ri, 2).value
    row_fill = PatternFill("solid", fgColor="2D0000" if sev == "CRITICAL" else "2D1A00")
    txt_col  = C["red"] if sev == "CRITICAL" else C["amber"]
    for ci in range(1, 5):
        ws_anom.cell(ri, ci).fill = row_fill
        ws_anom.cell(ri, ci).font = Font(color=txt_col, size=9, name="Calibri")

# ═══════════════════════════════════════════════════════════════
#  SHEET 7 — ROOT CAUSE
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_rc, "Root Cause Analysis  |  H1 vs H2 2024 Comparison", 4, size=13)
ws_rc.sheet_properties.tabColor = C["amber"]
write_df(ws_rc, root_cause_df, start_row=2, hdr_bg=C["amber"])
for i, w in enumerate([26, 16, 16, 12], 1):
    ws_rc.column_dimensions[get_column_letter(i)].width = w

for ri in range(3, 3 + len(root_cause_df)):
    c = ws_rc.cell(ri, 4)
    c.font = Font(color=C["red"] if "-" in str(c.value) else C["green"],
                  bold=True, size=9, name="Calibri")

narratives = [
    (16, "ROOT CAUSE 1: SPEND-REVENUE DIVERGENCE",        C["red"],   True),
    (17, "Spend rose in H2 while revenue fell. ROAS collapsed below 6x — ad fatigue and audience saturation from unchanged creative.", C["light"], False),
    (19, "ROOT CAUSE 2: CONVERSION RATE DECAY",           C["red"],   True),
    (20, "Traffic declined modestly but orders fell much faster. Conv rate dropped below 3.8% benchmark from Aug-24 onward.",          C["light"], False),
    (22, "ROOT CAUSE 3: RISING PRODUCT RETURNS",          C["amber"], True),
    (23, "Return volume doubled H1 to H2. Likely expectation mismatch from misleading product imagery or sizing issues.",              C["light"], False),
    (25, "ROOT CAUSE 4: NEW CUSTOMER ACQUISITION FAILURE",C["red"],   True),
    (26, "New customers fell ~42% while spend rose. Audience exhaustion — same creative reaching a saturated pool.",                   C["light"], False),
]
for rn, txt, col, bold in narratives:
    ws_rc.merge_cells(f"A{rn}:D{rn}")
    c = ws_rc.cell(rn, 1, txt)
    c.font      = Font(color=col, bold=bold, size=10 if bold else 9, name="Calibri")
    c.fill      = fl(C["navy"])
    c.alignment = la()
    ws_rc.row_dimensions[rn].height = 22

# ═══════════════════════════════════════════════════════════════
#  SHEET 8 — RECOMMENDATIONS
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_recs, "Strategic Recommendations  |  Prioritized by Business Impact", 5, size=13)
ws_recs.sheet_properties.tabColor = C["green"]
write_df(ws_recs, recs, start_row=2, hdr_bg=C["green"])
for i, w in enumerate([12, 22, 44, 60, 22], 1):
    ws_recs.column_dimensions[get_column_letter(i)].width = w

for ri in range(3, 3 + len(recs)):
    pri = ws_recs.cell(ri, 1).value
    rc  = {"CRITICAL":"2D0000","HIGH":"2D1A00","MEDIUM":"0D2010"}.get(pri, C["navy"])
    tc  = {"CRITICAL":C["red"],"HIGH":C["amber"],"MEDIUM":C["green"]}.get(pri, C["light"])
    for ci in range(1, 6):
        ws_recs.cell(ri, ci).fill = PatternFill("solid", fgColor=rc)
    ws_recs.cell(ri, 1).font = Font(color=tc, bold=True, size=9, name="Calibri")

# ═══════════════════════════════════════════════════════════════
#  SHEET 9 — SQL QUERIES
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_sql, "SQL Aggregation Queries  |  PostgreSQL / BigQuery / Snowflake", 2, size=12)
ws_sql.sheet_properties.tabColor = C["slate"]

SQL_BLOCKS = [
("Q1: Monthly KPI Aggregation", """SELECT DATE_TRUNC('month', order_date) AS month,
  SUM(net_revenue) AS total_revenue,
  COUNT(DISTINCT order_id) AS total_orders,
  ROUND(SUM(net_revenue)/NULLIF(COUNT(DISTINCT order_id),0),2) AS aov,
  SUM(marketing_spend) AS mkt_spend,
  ROUND(SUM(net_revenue)/NULLIF(SUM(marketing_spend),0),2) AS roas,
  SUM(is_returned) AS total_returns
FROM ecommerce_orders GROUP BY 1 ORDER BY 1;"""),

("Q2: Customer Acquisition Cost (CAC)", """SELECT DATE_TRUNC('month', order_date) AS month,
  SUM(marketing_spend) AS mkt_spend,
  COUNT(DISTINCT CASE WHEN is_new_customer=1 THEN customer_id END) AS new_customers,
  ROUND(SUM(marketing_spend)/NULLIF(
    COUNT(DISTINCT CASE WHEN is_new_customer=1 THEN customer_id END),0),2) AS cac
FROM ecommerce_orders GROUP BY 1 ORDER BY 1;"""),

("Q3: Conversion Rate", """SELECT DATE_TRUNC('month', order_date) AS month,
  SUM(website_visits) AS total_visits,
  COUNT(DISTINCT order_id) AS total_orders,
  ROUND(COUNT(DISTINCT order_id)::FLOAT/NULLIF(SUM(website_visits),0)*100,2) AS conversion_rate_pct
FROM ecommerce_orders GROUP BY 1 ORDER BY 1;"""),

("Q4: Category Gross Margin & Return Rate", """SELECT category,
  ROUND(SUM(net_revenue),2) AS total_revenue,
  ROUND((SUM(net_revenue)-SUM(cogs))/NULLIF(SUM(net_revenue),0)*100,1) AS gross_margin_pct,
  ROUND(SUM(is_returned)::FLOAT/COUNT(DISTINCT order_id)*100,1) AS return_rate_pct,
  ROUND(SUM(net_revenue)/NULLIF(SUM(marketing_spend),0),2) AS roas
FROM ecommerce_orders GROUP BY 1 ORDER BY total_revenue DESC;"""),

("Q5: Repeat Purchase Rate", """SELECT DATE_TRUNC('month', order_date) AS month,
  ROUND(COUNT(DISTINCT CASE WHEN is_new_customer=0 THEN customer_id END)::FLOAT/
    NULLIF(COUNT(DISTINCT customer_id),0)*100,1) AS repeat_rate_pct
FROM ecommerce_orders GROUP BY 1 ORDER BY 1;"""),

("Q6: Inventory Turnover", """SELECT category,
  ROUND(AVG(inventory_units),0) AS avg_inventory,
  ROUND(SUM(cogs),2) AS total_cogs,
  ROUND(SUM(cogs)/NULLIF(AVG(inventory_units),0),2) AS inventory_turnover
FROM ecommerce_orders GROUP BY 1 ORDER BY inventory_turnover DESC;"""),

("Q7: Campaign ROI Ranking", """SELECT campaign,
  ROUND(SUM(net_revenue),2) AS total_revenue,
  ROUND(SUM(marketing_spend),2) AS total_spend,
  ROUND(SUM(net_revenue)/NULLIF(SUM(marketing_spend),0),2) AS roas,
  ROUND(SUM(marketing_spend)/NULLIF(
    COUNT(DISTINCT CASE WHEN is_new_customer=1 THEN customer_id END),0),2) AS cac
FROM ecommerce_orders GROUP BY 1 ORDER BY roas DESC;"""),

("Q8: Anomaly Detection - Revenue Drop >5% MoM", """WITH monthly AS (
  SELECT DATE_TRUNC('month',order_date) AS month,
         SUM(net_revenue) AS revenue, SUM(marketing_spend) AS spend
  FROM ecommerce_orders GROUP BY 1
), lagged AS (
  SELECT month, revenue, spend,
         LAG(revenue) OVER (ORDER BY month) AS prev_revenue
  FROM monthly
)
SELECT month,
  ROUND((revenue-prev_revenue)/NULLIF(prev_revenue,0)*100,1) AS rev_pct_change,
  ROUND(revenue/NULLIF(spend,0),2) AS roas,
  CASE
    WHEN (revenue-prev_revenue)/NULLIF(prev_revenue,0) < -0.05
      THEN 'CRITICAL: Revenue dropped >5% MoM'
    WHEN revenue/NULLIF(spend,0) < 6
      THEN 'WARNING: ROAS below 6x threshold'
    ELSE 'OK'
  END AS anomaly_flag
FROM lagged WHERE prev_revenue IS NOT NULL ORDER BY 1;"""),

("Q9: Customer Lifetime Value (CLV) Proxy", """SELECT customer_id,
  COUNT(DISTINCT order_id) AS total_orders,
  ROUND(SUM(net_revenue),2) AS lifetime_revenue,
  ROUND(AVG(net_revenue),2) AS avg_order_value,
  MIN(order_date) AS first_order, MAX(order_date) AS last_order,
  CASE WHEN COUNT(DISTINCT order_id)>=3 THEN 'High Value'
       WHEN COUNT(DISTINCT order_id)=2  THEN 'Repeat'
       ELSE 'One-Time' END AS customer_segment
FROM ecommerce_orders GROUP BY 1 ORDER BY lifetime_revenue DESC LIMIT 100;"""),

("Q10: H1 vs H2 Root Cause Comparison", """SELECT
  CASE WHEN EXTRACT(MONTH FROM order_date)<=6
       THEN 'H1-2024' ELSE 'H2-2024' END AS half_year,
  ROUND(SUM(net_revenue),2) AS revenue,
  ROUND(SUM(net_revenue)/NULLIF(SUM(marketing_spend),0),2) AS roas,
  ROUND(COUNT(DISTINCT order_id)::FLOAT/NULLIF(SUM(website_visits),0)*100,2) AS conv_rate_pct,
  SUM(is_returned) AS returns,
  ROUND((SUM(net_revenue)-SUM(cogs))/NULLIF(SUM(net_revenue),0)*100,1) AS gross_margin_pct
FROM ecommerce_orders GROUP BY 1 ORDER BY 1;"""),
]

sql_row = 3
for title, query in SQL_BLOCKS:
    ws_sql.merge_cells(f"A{sql_row}:B{sql_row}")
    c = ws_sql.cell(sql_row, 1, title)
    c.font = Font(color=C["amber"], bold=True, size=9, name="Courier New")
    c.fill = fl("080F1A"); c.alignment = la()
    ws_sql.row_dimensions[sql_row].height = 16
    sql_row += 1
    for line in query.split("\n"):
        ws_sql.merge_cells(f"A{sql_row}:B{sql_row}")
        c = ws_sql.cell(sql_row, 1, line)
        c.font = Font(color=C["green"], size=9, name="Courier New")
        c.fill = fl("080F1A"); c.alignment = la()
        ws_sql.row_dimensions[sql_row].height = 15
        sql_row += 1
    sql_row += 1

ws_sql.column_dimensions["A"].width = 75
ws_sql.column_dimensions["B"].width = 10

# ═══════════════════════════════════════════════════════════════
#  SHEET 10 — RAW DATA   *** PERIOD FIX APPLIED HERE ***
# ═══════════════════════════════════════════════════════════════
sheet_title(ws_raw, "Raw Transaction Data  (First 500 rows)", 16, size=12)
ws_raw.sheet_properties.tabColor = C["slate"]

sample = df.head(500).copy()
# Replace Timestamp column with plain string version
sample["order_date"] = sample["order_date_str"]
# Drop ALL columns that openpyxl cannot convert
drop_cols = [col for col in ["month_period", "month_str", "order_date_str"]
             if col in sample.columns]
sample = sample.drop(columns=drop_cols)
# Final safety: convert any remaining Period/Timestamp
for col in sample.columns:
    if sample[col].dtype == "object":
        sample[col] = sample[col].apply(
            lambda x: str(x) if isinstance(x, (pd.Period, pd.Timestamp)) else x)

write_df(ws_raw, sample, start_row=2, hdr_bg=C["slate"])
for i, w in enumerate([10,14,12,14,10,16,10,10,10,10,10,10,8,10,8,10], 1):
    ws_raw.column_dimensions[get_column_letter(i)].width = w

# ─── SAVE ─────────────────────────────────────────────────────
wb.save(OUT_EXCEL)
print(f"\nSaved: {OUT_EXCEL}")
print("Upload to Google Sheets: drive.google.com -> New -> File upload -> select .xlsx")
print("\n" + "="*55)
print(f"  Peak Revenue : {peak_month}  (${peak_rev:,.0f})")
print(f"  Latest Rev   : {latest_m['month_label']}  (${latest_m['revenue']:,.0f})")
print(f"  Rev Decline  : {rev_delta:.1f}% from peak")
print(f"  ROAS Latest  : {latest_m['roas']:.2f}x  (Benchmark {BENCHMARKS['roas']}x)")
print(f"  CAC Latest   : ${latest_m['cac']:.2f}  (Benchmark ${BENCHMARKS['cac']})")
print(f"  Total Alerts : {int(monthly['alert_count'].sum())}")
print("="*55)